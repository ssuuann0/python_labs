from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv 

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    p_xlsx=Path(xlsx_path)
    p_csv=Path(csv_path)
    if p_xlsx.suffix.lower() != '.xlsx' or p_csv.suffix.lower() != '.csv': #проверка типа файла
        raise ValueError ('Неверный тип файла')
    wb = Workbook() #создает таблицу 
    ws = wb.active #создает пустой лист
    ws.title = "Sheet1" #дает листу имя
    try:
        with open(csv_path, encoding="utf-8") as f:
            csv_dict=list(csv.DictReader(csv_path))
            if len(csv_dict)==0:
                raise ValueError ('Пустой файл')
            k_str=len(csv_dict[0])
            col_width=8
            for col in range(1,k_str+1):
                col_letter=get_column_letter(col) #преобразует номер столбца в букву(как в Exel)
                ws.column_dimensions[col_letter].width=col_width #установление ширины колонки
            for row in csv.reader(f):
                ws.append(row)
            wb.save(xlsx_path)
    except:
        raise FileNotFoundError ('Файл не найден')
csv_to_xlsx(r'C:\Users\Соня\OneDrive\Рабочий стол\git\python_labs-1\src\data\lab_05\samples\people.csv',r'C:\Users\Соня\OneDrive\Рабочий стол\git\python_labs-1\src\data\lab_05\out\people.xlsx')
